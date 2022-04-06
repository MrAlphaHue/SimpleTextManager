# Python program to read an excel file
# import openpyxl module
import time
import openpyxl
import ConstManage as ConstManage
import glob
import base64
import os;

#Avaliable Type :  Xint , int , string , Enum? , double , float

#print  (glob.glob('/Users/munjeonghwan/OneFolder/UnityProject_Launched/Utils_SimpleDataConverter/Xlsxs/*.xlsx'))
#print (os.getcwd())

start = time.time()  # 시작 시간 저장
xlsxPath = ConstManage.xlsx_path + "/*.xlsx";

def makeBinaryByXlsx(work_book_structs):
    print("Execute Make BinaryByXlsx")

    files =  glob.glob(ConstManage.binary_path + "/*.bytes")
    for f in files:
        print("Delete " + f)
        os.remove(f)

    files =  glob.glob(ConstManage.project_binary_path + "/*.bytes")
    for f in files:
        print("Delete " + f)
        os.remove(f)

    for wbStructs in work_book_structs:
        for sheet_struct in wbStructs.sheet_structs:
            file_name = wbStructs.base_filename + sheet_struct.sheet.title;
            text_to_file = ""

            #1열은 변수 타입 , 2열은 변수명.
            for idx_row in range(1, sheet_struct.max_row + 1):
                for idx_column in range(1, sheet_struct.max_column + 1):
                    cell_obj = sheet_struct.sheet.cell(row = idx_row, column = idx_column)
                    if not (idx_row == 1 and idx_column == 1):
                        text_to_file += '\t'
                    text_to_file += str(cell_obj.value)
                if not idx_row == sheet_struct.max_row:
                    text_to_file += '\t'


            ##
            utf8_encode = text_to_file.encode() 
            utf8_encode_base = base64.encodebytes(utf8_encode)

            ##
            file = open(ConstManage.binary_path + "/" + file_name + ".bytes", "wb")
            file.write(utf8_encode_base)

            ##
            file_proejct = open(ConstManage.project_binary_path + "/" + file_name + ".bytes", "wb")
            file_proejct.write(utf8_encode_base)
    

print("time :", time.time() - start)  # 현재시각 - 시작시간 = 실행 시간