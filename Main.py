
# 1. 모든 워크시트를 한 리스트에 넌다. 
# 2. 해당 테이블들 cell 값에 tab이 없는지 체크한다.
# 3. 




class WorkSheetStruct:
    def __init__(self, sheet, max_row, max_column):
        self.sheet = sheet
        self.max_row = max_row
        self.max_column = max_column

class WorkBookStruct:
    def __init__(self, work_book, base_filename):
        self.work_book = work_book
        self.base_filename = base_filename
        self.sheet_structs = []
        allsheet = work_book.worksheets;
        for sheet_obj in allsheet:
            if("TEMP" in sheet_obj.title.upper()):
                continue
            max_row = sheet_obj.max_row
            max_column = sheet_obj.max_column
            for idx_row in range(1, max_row + 1):
                #id 란이 공백일 경우 (column = 1 means 'id')
                #id . 0 , 1, 2, 3,
                print(str(sheet_obj.cell(row = idx_row , column = 1).value))
                if(sheet_obj.cell(row = idx_row , column = 1).value is None):
                    max_row = idx_row - 1;
                    break
                #id, strKey , en, kr
            for idx_column in range(1, max_column + 1):
                print(str(sheet_obj.cell(row = 1 , column = idx_column).value))
                if(sheet_obj.cell(row = 1 , column = idx_column).value is None):
                    max_column = idx_column - 1;
                    break    
            sheet_struct = WorkSheetStruct(sheet = sheet_obj, max_row = max_row, max_column = max_column)
            self.sheet_structs.append(sheet_struct)


# Python program to read an excel file
# import openpyxl module
import time
# from typing import List
import openpyxl
import glob
import base64
import os

#Custom
import ConstManage as ConstManage
import FunctionContainer as Func
import MakeBinaryFile
import MakeScriptFile
import MakeGeneralScriptFile

#Avaliable Type :  Xint , int , string , Enum? , double , float

#print  (glob.glob('/Users/munjeonghwan/OneFolder/UnityProject_Launched/Utils_SimpleDataConverter/Xlsxs/*.xlsx'))
#print  (os.getcwd())


start = time.time()  # 시작 시간 저장
xlsxPath = ConstManage.xlsx_path + "/*.xlsx";

# def GetTableSheetLists()


##
# MakeGeneralScriptFile.Generate()
workBookStructs = []

##
def MakeWorkBookStructs():
    ##
    workBookStructs.clear()
    ##
    for filename in glob.glob(xlsxPath):
        if ("~$" in filename):
            continue
        # Must have "TexTable" in Name
        if("TexTable" not in filename):
            continue

        wb_obj = openpyxl.load_workbook(filename, read_only=False, data_only=True) #read_only = true, really loosy.
        base_filename = Func.func_getBasefileName(filename)

        wb_struct = WorkBookStruct(work_book = wb_obj, base_filename = base_filename)
        workBookStructs.append(wb_struct)

    ##
    print("We Have " + str(len(workBookStructs)) + " Workbooks")
    
    ##
    for wbStr in workBookStructs:
        print("WorkBook Name " + wbStr.base_filename)
        for sheetStr in wbStr.sheet_structs:
            print("sheetName " + sheetStr.sheet.title)



def CheckTabInCell(work_book_structs):
    flg = True;

    for wbStructs in work_book_structs:
        allsheet = wbStructs.work_book.worksheets;

        for sheet_obj in allsheet:
            if("TEMP" in sheet_obj.title.upper()):
                continue

            m_row = sheet_obj.max_row
            m_column = sheet_obj.max_column

            for i in range(1, m_row + 1):
                for j in range(1, m_column + 1):
                    cell_obj = sheet_obj.cell(row = i, column = j)
                    if(cell_obj.value is None):
                        continue
                    if('\t' in str(cell_obj.value)):
                        print("Has Tab at  \"" + wbStructs.base_filename + " : [" + Func.func_intToAlphabet(j) + str(i) +  "]\" !!!!!!!!")
                        flg = False
    return flg

def Main():
    print("Progress 0 Execute\n")
    MakeWorkBookStructs()

    print("Progress 1 Execute\n")
    if(CheckTabInCell(workBookStructs) is False):
        return
    
    MakeBinaryFile.makeBinaryByXlsx(workBookStructs)
    MakeScriptFile.makeScriptByXlsx(workBookStructs)



Main()

    



print("End time :", time.time() - start)  # 현재시각 -  시작시간 = 실행 시간